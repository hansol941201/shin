#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
POUR 공법 실적 List 엑셀(연도별 시트)을 앱 실적 자료로 옮긴다.

- 시트 이름의 연도를 그대로 record.year 로 쓴다. (원본에 날짜가 없으므로 날짜는 만들어 내지 않는다)
- 특허번호 칸의 숫자는 POUR 특허 항목으로, 숫자가 없는 표기("POUR공법" 등)는
  공고문 특허·공법 원문(noticePatentText)에 그대로 남긴다.
- 같은 현장·같은 공사가 여러 시트에 반복되면 한 행으로 합친다. 값이 빈 칸만 채워 넣으므로
  어느 시트의 정보도 사라지지 않는다.

사용: python3 scripts/import-records.py <엑셀경로> [출력 JSON] [출력 SQL]
"""
import json
import re
import sys
import unicodedata

import openpyxl

HEADER_ROW = 4
COL = {                       # 1-기준 열 번호
    "seq": 1, "categories": 2, "region": 3, "city": 4, "patent": 5,
    "client": 6, "project": 7, "scope": 8, "phone": 9, "households": 10,
    "quality": 11, "contractor": 12, "note": 15,
}
LAST_COL = 16


def text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        value = int(value)
    return unicodedata.normalize("NFC", str(value)).strip()


def one_line(value):
    """한 줄짜리 칸에 들어온 줄바꿈·중복 공백을 정리한다. 내용은 지우지 않는다."""
    return re.sub(r"\s+", " ", text(value)).strip()


def to_list(value):
    """줄바꿈·쉼표로 나뉜 칸을 목록으로 만든다. 순서와 중복 제거를 함께 처리한다."""
    raw = text(value)
    if not raw:
        return []
    parts = [p.strip(" \t·-") for p in re.split(r"[\n\r,]+", raw)]
    out, seen = [], set()
    for p in parts:
        if p and p not in seen:
            seen.add(p)
            out.append(p)
    return out


def phone_text(value):
    """앞자리 0이 사라지지 않도록 문자열로 두고 공백만 정리한다."""
    raw = text(value)
    if not raw:
        return ""
    raw = re.sub(r"\s*-\s*", "-", raw)
    return re.sub(r"\s+", " ", raw).strip()


def digits(value):
    return re.sub(r"[^0-9]", "", text(value))


# 특허번호가 아닌 표기. 숫자가 섞여 있어도 특허번호로 바꾸지 않는다.
NOT_A_PATENT = re.compile(r"신기술|공법|파우더|인증|규격")


def normalize_patent_number(raw):
    """pour-patents.js 의 normalizeNumber 와 같은 규칙에 원본 검사만 더한다."""
    text_value = text(raw)
    if NOT_A_PATENT.search(text_value):
        return ""                       # "건설신기술 1026호" 같은 표기는 번호가 아니다
    d = digits(text_value)
    if not d:
        return ""
    if len(d) in (9, 13) and (d.startswith("10") or d.startswith("20")):
        d = d[2:]
    # 등록번호는 여섯 자리보다 짧을 수 없다. 짧으면 번호가 아닌 다른 표기로 본다.
    return d if len(d) >= 6 else ""


def format_patent_number(number):
    return "제10-%s호" % number if number else ""


def households(value):
    d = digits(value)
    return int(d) if d else ""


# 공종 분류표 — pour-categories.js 와 같은 표를 쓴다
CATEGORY_GROUPS = [
    ("옥상방수", ["PVC", "금속기와", "박공지붕", "방수", "복합시트", "슬라브", "싱글", "우레탄"]),
    ("재도장",   ["균열보수", "재도장"]),
    ("주차장",   ["균열보수", "배면차수", "아스콘", "에폭시", "우레탄", "재도장"]),
    ("도로",     ["보도블럭", "아스콘", "에폭시"]),
]
CATEGORY_OTHER = "기타"


def classify_category(name):
    """대분류가 하나로 정해질 때만 분류한다. 확실하지 않으면 기타로 두고 이름은 그대로."""
    clean = text(name)
    if not clean:
        return None
    plain = re.sub(r"[\s·]", "", clean).upper()
    found = [(group, item) for group, items in CATEGORY_GROUPS for item in items
             if re.sub(r"[\s·]", "", item).upper() == plain]
    if len(found) == 1:
        return {"group": found[0][0], "name": found[0][1]}
    return {"group": CATEGORY_OTHER, "name": clean}


def classify_categories(names):
    out, seen = [], set()
    for name in names:
        item = classify_category(name)
        if not item:
            continue
        k = (item["group"], re.sub(r"[\s·]", "", item["name"]).upper())
        if k in seen:
            continue
        seen.add(k)
        out.append(item)
    return out


def key_text(value):
    """중복 판단용 열쇠. 공백·괄호·기호를 지우고 비교한다."""
    return re.sub(r"[\s()（）\[\]·,./\-]", "", text(value))


def is_empty_row(row):
    """지역·도시만 남고 알맹이가 없는 줄. 실적으로 볼 수 없으므로 세지 않는다."""
    return not (row["client"] or row["projectNames"] or row["patentNumbers"]
                or row["noticePatentText"] or digits(row["phone"]))


def read_sheet(ws, year):
    rows = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        cells = [ws.cell(r, c).value for c in range(1, LAST_COL + 1)]
        if not any(text(v) for v in cells[1:]):     # 순번만 있는 빈 줄은 건너뛴다
            continue

        patent_numbers, patent_texts = [], []
        for part in to_list(cells[COL["patent"] - 1]):
            number = normalize_patent_number(part)
            if number:
                if number not in patent_numbers:
                    patent_numbers.append(number)
            elif part not in patent_texts:
                patent_texts.append(part)

        row = {
            "year": year,
            "sheet": ws.title,
            "sourceRow": r,
            "categories": to_list(cells[COL["categories"] - 1]),
            "region": one_line(cells[COL["region"] - 1]),
            "city": one_line(cells[COL["city"] - 1]),
            "patentNumbers": patent_numbers,
            "noticePatentText": "\n".join(patent_texts),
            "client": one_line(cells[COL["client"] - 1]),
            "projectNames": to_list(cells[COL["project"] - 1]),
            "scopes": to_list(cells[COL["scope"] - 1]),
            "phone": phone_text(cells[COL["phone"] - 1]),
            "households": households(cells[COL["households"] - 1]),
            "quality": one_line(cells[COL["quality"] - 1]),
            "contractor": one_line(cells[COL["contractor"] - 1]),
            "remark": text(cells[COL["note"] - 1]),
        }
        if not is_empty_row(row):
            rows.append(row)
    return rows


def strict_key(row):
    """같은 현장·같은 공사·같은 전화·같은 공종·같은 특허 = 같은 실적."""
    return (
        key_text(row["client"]),
        key_text(" ".join(row["projectNames"])),
        digits(row["phone"]),
        key_text(" ".join(row["categories"])),
        ",".join(sorted(row["patentNumbers"])),
    )


def loose_key(row):
    """전화·공종·특허 표기만 다른 겹침도 걸러 내려고 현장+공사명으로 한 번 더 본다."""
    return (key_text(row["client"]), key_text(" ".join(row["projectNames"])))


LIST_FIELDS = ("categories", "projectNames", "scopes", "patentNumbers")
TEXT_FIELDS = ("region", "city", "client", "phone", "quality", "contractor",
               "remark", "noticePatentText")


def merge(base, other):
    """빈 칸만 채운다. 이미 값이 있는 칸은 덮어쓰지 않는다."""
    for f in TEXT_FIELDS:
        if not base[f] and other[f]:
            base[f] = other[f]
    if base["households"] == "" and other["households"] != "":
        base["households"] = other["households"]
    for f in LIST_FIELDS:
        for v in other[f]:
            if v not in base[f]:
                base[f].append(v)
    for y in other["years"]:
        if y not in base["years"]:
            base["years"].append(y)
    base["years"].sort(reverse=True)
    base["year"] = base["years"][0]


def to_record(row, index):
    stamp = "%s-01-01 00:00:00" % (row["year"] or "2018")
    items = []
    for i, number in enumerate(row["patentNumbers"]):
        items.append({
            "id": "pat-imp-%04d-%d" % (index, i),
            "kind": "POUR",
            "number": number,
            "display": format_patent_number(number),
            "name": "", "method": "", "company": "", "category": "", "remark": "",
            "createdAt": stamp, "updatedAt": stamp,
        })
    category_items = classify_categories(row["categories"])
    years = row["years"]
    remark = row["remark"]
    if len(years) > 1:
        note = "실적 List 등재 연도: " + ", ".join("%s년" % y for y in years)
        remark = (remark + "\n" + note).strip() if remark else note
    return {
        "id": "rec-imp-%04d" % index,
        # 엑셀에서 옮겨 온 행 표시. "협약서번호 미입력" 알림 대상에서 빼는 데만 쓴다.
        "source": "import",
        "status": "낙찰",
        "year": row["year"],
        "categories": row["categories"],
        "categoryItems": category_items,
        "categoryGroups": sorted(
            {it["group"] for it in category_items},
            key=lambda g: [x[0] for x in CATEGORY_GROUPS].index(g)
            if g in [x[0] for x in CATEGORY_GROUPS] else 99),
        "region": row["region"],
        "city": row["city"],
        "patentItems": items,
        "noticePatentText": row["noticePatentText"],
        "client": row["client"],
        "projectNames": row["projectNames"],
        "scopes": row["scopes"],
        "phone": row["phone"],
        "households": row["households"],
        "quality": row["quality"],
        "contractor": row["contractor"],
        "remark": remark,
        "createdAt": stamp,
    }


def q(value):
    """SQL 문자열 리터럴. 전화번호는 앞자리 0이 살아 있어야 하므로 언제나 문자열이다."""
    if value is None or value == "":
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def build_sql(records):
    """D1 에 옮기는 SQL. 지우는 문장은 쓰지 않고, 같은 id 는 덮어쓴다."""
    head = [
        "-- 연도별 POUR 실적 List 엑셀 → projects / pour_project_patents",
        "-- scripts/import-records.py 가 만든 파일입니다. 직접 고치지 마세요.",
        "--",
        "-- 원칙",
        "--   · DELETE / DROP / TRUNCATE 를 쓰지 않습니다 (기존 운영 자료 보존)",
        "--   · 같은 id 로 다시 실행해도 행이 늘지 않습니다 (UPSERT)",
        "--   · 0002~0005 마이그레이션(record_year·category_items·record_source 포함)을 먼저 실행하세요",
        "",
        "BEGIN TRANSACTION;",
        "",
    ]
    body = []
    for rec in records:
        cols = [
            ("id", rec["id"]), ("status", rec["status"]), ("record_year", rec["year"]),
            # 엑셀 이전분 표시. 협약서번호 미입력 알림에서 빼는 데만 쓴다.
            ("record_source", rec.get("source") or None),
            ("region", rec["region"]), ("city", rec["city"]), ("client", rec["client"]),
            ("project_name", "\n".join(rec["projectNames"])),
            ("category", "\n".join(rec["categories"])),
            ("category_items", json.dumps(rec["categoryItems"], ensure_ascii=False)
             if rec["categoryItems"] else None),
            ("scopes", "\n".join(rec["scopes"])),
            ("phone", rec["phone"]),
            ("households", rec["households"] if rec["households"] != "" else None),
            ("quality", rec["quality"]), ("contractor", rec["contractor"]),
            ("notice_patent_text", rec["noticePatentText"]), ("remark", rec["remark"]),
            ("patents_migrated", 1),
            ("created_at", rec["createdAt"]), ("updated_at", rec["createdAt"]),
        ]
        names = [c[0] for c in cols]
        values = []
        for name, v in cols:
            values.append(str(v) if name in ("households", "patents_migrated") and v is not None
                          else q(v))
        updates = ", ".join("%s=excluded.%s" % (n, n) for n in names if n != "id")
        body.append("INSERT INTO projects (%s) VALUES (%s)\n  ON CONFLICT(id) DO UPDATE SET %s;"
                    % (", ".join(names), ", ".join(values), updates))
        for order, item in enumerate(rec["patentItems"]):
            body.append(
                "INSERT INTO pour_project_patents "
                "(id, project_id, kind, number, display, sort_order, created_at, updated_at)\n"
                "  VALUES (%s, %s, 'POUR', %s, %s, %d, %s, %s)\n"
                "  ON CONFLICT(project_id, kind, number) DO UPDATE SET "
                "display=excluded.display, sort_order=excluded.sort_order, "
                "updated_at=excluded.updated_at;"
                % (q(item["id"]), q(rec["id"]), q(item["number"]), q(item["display"]),
                   order, q(item["createdAt"]), q(item["updatedAt"])))
    return "\n".join(head + body + ["", "COMMIT;", ""])


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else ""
    out = sys.argv[2] if len(sys.argv) > 2 else "test/fixtures-records.json"
    wb = openpyxl.load_workbook(src, data_only=True)

    raw = []
    per_sheet = []
    for name in wb.sheetnames:
        year = (re.search(r"(20\d{2})", name) or [None, ""])[1]
        rows = read_sheet(wb[name], year)
        per_sheet.append((name, len(rows)))
        raw.extend(rows)

    for row in raw:
        row["years"] = [row["year"]] if row["year"] else []

    duplicates = 0

    def collapse(rows, key_of):
        nonlocal duplicates
        table, order = {}, []
        for row in rows:
            k = key_of(row)
            if k in table:
                duplicates += 1
                merge(table[k], row)
            else:
                table[k] = row
                order.append(k)
        return [table[k] for k in order]

    rows = collapse(raw, strict_key)          # 완전히 같은 줄부터 합치고
    rows = collapse(rows, loose_key)          # 표기만 다른 겹침을 한 번 더 합친다

    records = [to_record(row, i + 1) for i, row in enumerate(rows)]

    with open(out, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=1)

    sql_out = sys.argv[3] if len(sys.argv) > 3 else "nextjs/drizzle/seed-records.sql"
    with open(sql_out, "w", encoding="utf-8") as f:
        f.write(build_sql(records))

    for name, n in per_sheet:
        print("  %-12s %4d행" % (name, n))
    print("원본 합계 %d행 / 중복 병합 %d행 / 최종 %d행" % (len(raw), duplicates, len(records)))
    print("→ %s" % out)
    print("→ %s" % sql_out)


if __name__ == "__main__":
    main()

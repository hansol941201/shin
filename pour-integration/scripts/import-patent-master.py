#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
월말실적 엑셀에서 특허 마스터와 현장별 특허 기록을 뽑는다.

  python3 pour-integration/scripts/import-patent-master.py <월말실적.xlsx>

읽는 시트
  「특허(N)」          특허 마스터  — 특허번호 / 업체명 / 공법명 / 발명의 명칭
  「26년 피벗용(쿼리)」 현장별 기록  — 단지명 / 특허번호 / 업체명 / 공종 / 지역 / 공고일

만드는 것
  test/fixtures-patent-master.json   특허 마스터 (번호마다 한 건)
  nextjs/drizzle/seed-patent-master.sql  같은 내용의 UPSERT SQL

원칙
  · 업체명·공법명을 추정하지 않는다. 엑셀에 없으면 비워 두고 "미분류" 로 남긴다
  · 같은 특허번호는 한 건으로 합친다 (중복 생성 금지)
  · 개별 특허의 구분과 현장 전체의 구분을 섞지 않는다
    - 엑셀의 "구분" 열(타사·다특허·POUR·다특허(PD)·DO·CNC)은 그 현장의 구분이다
    - 개별 특허의 구분은 업체명이 POUR 업체인지로만 정하고, 모르면 미분류로 둔다
  · DELETE / DROP / 표 교체를 쓰지 않는다
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent

# ---------------------------------------------------------------- 분류 기준
#
# 소속과 공법을 나누어 담는다. DO 와 CNC 는 자사 계열이지만 POUR 공법은 아니다.
#
#   소속  자사계열 / 타사 / 미분류
#   공법  POUR / DO / CNC / 타사공법 / 미분류
#
# 아래 이름들은 모두 원본 엑셀에 실제로 적혀 있는 값이다. 짐작으로 만든 것이 없다.
# 여기 없는 업체는 추정하지 않고 타사(업체명이 있을 때) 또는 미분류로 둔다.

# 「특허(N)」 업체명 열에 그대로 적혀 있는 이름
POUR_COMPANIES = {"㈜넷폼알앤디", "넷폼알앤디", "(주)넷폼알앤디"}
DO_COMPANIES = {"DO공법"}
CNC_COMPANIES = {"CNC공법"}

AFF_OWN, AFF_THIRD, AFF_UNKNOWN = "자사계열", "타사", "미분류"
TYPE_POUR, TYPE_DO, TYPE_CNC = "POUR", "DO", "CNC"
TYPE_THIRD, TYPE_UNKNOWN = "타사공법", "미분류"

AFFILIATION_BY_TYPE = {
    TYPE_POUR: AFF_OWN, TYPE_DO: AFF_OWN, TYPE_CNC: AFF_OWN,
    TYPE_THIRD: AFF_THIRD, TYPE_UNKNOWN: AFF_UNKNOWN,
}


def norm_number(raw):
    """
    특허번호에서 숫자만 남긴다. pour-patents.js 의 normalizeNumber 와 같은 규칙이라야
    앱이 쓰는 번호와 어긋나지 않는다.

      9자리  10-1935719     → 앞 두 자리(권리구분)를 뗀다
      13자리 10-2017-0012345 → 마찬가지
      그 밖  그대로 둔다

    「특허(N)」 시트 1열에는 공법명 글자도 섞여 있어("아크릴 2중주입공법"),
    숫자가 몇 개 안 나오는 것은 특허번호로 보지 않는다.
    """
    digits = re.sub(r"\D", "", str(raw or ""))
    if not digits:
        return ""
    if len(digits) in (9, 13) and (digits.startswith("10") or digits.startswith("20")):
        digits = digits[2:]
    # 등록번호는 7자리다. 그보다 짧으면 특허번호가 아니라 글자에서 딸려 나온 숫자다.
    if len(digits) < 7:
        return ""
    return digits


def text(v):
    return str(v).strip() if v is not None else ""


def patent_type_for(company, method=""):
    """
    개별 특허 한 건의 공법. 업체명을 모르면 미분류로 둔다 (추정하지 않는다).

    판정 근거는 원본 엑셀 두 곳이 서로 맞는 것만 쓴다.
      · 업체명이 "DO공법"  → DO   (「특허(N)」 7건, 피벗 시트 구분=DO 행의 번호와 일치)
      · 업체명이 "CNC공법" → CNC  (「특허(N)」 2건, 피벗 시트 구분=CNC 행과 일치)
      · 넷폼알앤디이면서 공법명에 CNC 가 적힌 것 → CNC (자사 계열의 CNC 공법)
      · 그 밖의 넷폼알앤디 → POUR
      · 업체명이 있으나 위에 없음 → 타사공법
      · 업체명 없음 → 미분류

    "다특허(PD)" 같은 값은 현장 전체의 구분이지 개별 특허의 공법이 아니므로 쓰지 않는다.
    """
    if not company:
        return TYPE_UNKNOWN
    if company in DO_COMPANIES:
        return TYPE_DO
    if company in CNC_COMPANIES:
        return TYPE_CNC
    if company in POUR_COMPANIES:
        # 같은 회사 안에서도 CNC 공법은 따로 센다
        return TYPE_CNC if "CNC" in method.upper() else TYPE_POUR
    return TYPE_THIRD


def affiliation_for(patent_type):
    return AFFILIATION_BY_TYPE.get(patent_type, AFF_UNKNOWN)


def read_master(wb):
    """「특허(N)」 시트 → 특허번호마다 한 건."""
    if "특허(N)" not in wb.sheetnames:
        return {}
    ws = wb["특허(N)"]
    # 제목 줄 찾기 (앞쪽 몇 줄 안에 "특허번호" 가 있다)
    header = None
    for r in range(1, min(ws.max_row, 10) + 1):
        if text(ws.cell(r, 1).value) == "특허번호":
            header = r
            break
    if header is None:
        return {}

    out = {}
    for r in range(header + 1, ws.max_row + 1):
        number = norm_number(ws.cell(r, 1).value)
        if not number:
            continue
        company = text(ws.cell(r, 3).value)
        method = text(ws.cell(r, 4).value)
        name = text(ws.cell(r, 5).value)     # 발명의 명칭
        if number in out:
            # 같은 번호가 여러 줄이면 비어 있는 칸만 채운다
            rec = out[number]
            if not rec["methodName"] and method:
                rec["methodName"] = method
            if not rec["company"] and company:
                rec["company"] = company
            if not rec["name"] and name:
                rec["name"] = name
            rec["patentType"] = patent_type_for(rec["company"], rec["methodName"])
            rec["affiliationType"] = affiliation_for(rec["patentType"])
            continue
        ptype = patent_type_for(company, method)
        out[number] = {
            "number": number,
            "name": name,
            "company": company,
            "methodName": method,
            "patentType": ptype,
            "affiliationType": affiliation_for(ptype),
            "categories": [],
            "category": "",
            "remark": "",
            "prefix": "",
            "active": True,
            "firstSeenAt": "",
            "lastSeenAt": "",
        }
    return out


def apply_single_site_evidence(master, sites):
    """
    특허 한 건짜리 현장의 "구분" 열을 근거로 쓴다.

    특허가 하나뿐인 현장의 구분은 곧 그 특허의 공법이므로 확실한 근거다.
    다만 구분이 "DO"·"CNC" 일 때만 쓴다. 원본의 "POUR" 는 자사 계열을 POUR 로
    묶어 세던 옛 방식이 섞여 있어(업체명이 DO공법인데 구분이 POUR 인 행이 있다)
    근거로 쓰지 않는다. 우리가 지금 나누려는 것이 바로 그 묶음이다.

    업체명으로 이미 정해진 것은 덮지 않는다. 미분류로 남은 것만 채운다.
    """
    per_site = {}
    for row in sites:
        per_site.setdefault(row["client"], set()).add(row["number"])

    filled = []
    for row in sites:
        if len(per_site[row["client"]]) != 1:
            continue
        raw = row["siteClassRaw"]
        if raw not in (TYPE_DO, TYPE_CNC):
            continue
        rec = master.get(row["number"])
        if not rec or rec["patentType"] != TYPE_UNKNOWN:
            continue                      # 업체명으로 이미 정해진 것은 그대로 둔다
        rec["patentType"] = raw
        rec["affiliationType"] = affiliation_for(raw)
        filled.append((row["number"], row["client"], raw))
    return filled


def read_sites(wb, master):
    """「26년 피벗용(쿼리)」 시트 → 현장별 특허 기록. 마스터에 없는 번호는 미분류로 더한다."""
    if "26년 피벗용(쿼리)" not in wb.sheetnames:
        return []
    ws = wb["26년 피벗용(쿼리)"]
    rows = []
    for r in range(2, ws.max_row + 1):
        client = text(ws.cell(r, 3).value)
        number = norm_number(ws.cell(r, 6).value)
        if not client or not number:
            continue
        company = text(ws.cell(r, 7).value)
        notice = ws.cell(r, 9).value
        rows.append({
            "client": client,
            "number": number,
            "company": company,
            # 엑셀의 "구분" 은 그 현장의 구분이다. 개별 특허 구분으로 쓰지 않는다.
            "siteClassRaw": text(ws.cell(r, 2).value),
            "categoryGroup": text(ws.cell(r, 4).value),
            "category": text(ws.cell(r, 5).value),
            "projectName": text(ws.cell(r, 8).value),
            "noticeDate": notice.strftime("%Y-%m-%d") if hasattr(notice, "strftime") else "",
            "region": text(ws.cell(r, 11).value),
            "city": text(ws.cell(r, 12).value),
        })
        # 마스터에 없는 번호는 여기서 미분류로 등록한다 (업체명은 엑셀에 있을 때만 담는다)
        if number not in master:
            ptype = patent_type_for(company)
            master[number] = {
                "number": number, "name": "", "company": company,
                "methodName": "", "patentType": ptype,
                "affiliationType": affiliation_for(ptype),
                "categories": [], "category": "", "remark": "", "prefix": "",
                "active": True, "firstSeenAt": "", "lastSeenAt": "",
            }
        elif not master[number]["company"] and company:
            master[number]["company"] = company
            master[number]["patentType"] = patent_type_for(company, master[number]["methodName"])
            master[number]["affiliationType"] = affiliation_for(master[number]["patentType"])
    return rows


def sql_text(v):
    if v is None or v == "":
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"파일을 찾지 못했습니다: {path}")
        return 1

    wb = openpyxl.load_workbook(path, data_only=True)
    master = read_master(wb)
    from_master = len(master)
    sites = read_sites(wb, master)
    filled = apply_single_site_evidence(master, sites)

    # 현장에서 처음/마지막으로 본 날을 채운다 (공고일 기준. 없으면 비워 둔다)
    seen = {}
    for row in sites:
        if not row["noticeDate"]:
            continue
        at = seen.setdefault(row["number"], [row["noticeDate"], row["noticeDate"]])
        at[0] = min(at[0], row["noticeDate"])
        at[1] = max(at[1], row["noticeDate"])
    for number, (first, last) in seen.items():
        master[number]["firstSeenAt"] = first
        master[number]["lastSeenAt"] = last

    records = sorted(master.values(), key=lambda r: r["number"])

    out_json = BASE / "test" / "fixtures-patent-master.json"
    out_json.write_text(json.dumps(
        {"master": records, "siteLinks": sites}, ensure_ascii=False, indent=1), encoding="utf-8")

    lines = [
        "-- 특허 마스터 옮겨 심기 (번호 기준 UPSERT · 지우지 않음)",
        "-- import-patent-master.py 가 만든 파일입니다. 손으로 고치지 마세요.",
        "-- 업체명·공법명이 비어 있는 건은 '미분류' 입니다. 값을 지어내지 않았습니다.",
        "",
    ]
    for rec in records:
        lines.append(
            "INSERT INTO pour_patents "
            "(number, display, name, categories, company, prefix, remark, active, "
            "created_at, updated_at, patent_type, method_name, first_seen_at, last_seen_at, "
            "affiliation_type) VALUES ("
            f"{sql_text(rec['number'])}, {sql_text('제10-' + rec['number'] + '호')}, "
            f"{sql_text(rec['name'])}, NULL, {sql_text(rec['company'])}, NULL, NULL, 1, "
            f"NULL, NULL, {sql_text(rec['patentType'])}, {sql_text(rec['methodName'])}, "
            f"{sql_text(rec['firstSeenAt'])}, {sql_text(rec['lastSeenAt'])}, "
            f"{sql_text(rec['affiliationType'])})\n"
            "ON CONFLICT(number) DO UPDATE SET\n"
            "  name = COALESCE(NULLIF(excluded.name, ''), pour_patents.name),\n"
            "  company = COALESCE(NULLIF(excluded.company, ''), pour_patents.company),\n"
            "  method_name = COALESCE(NULLIF(excluded.method_name, ''), pour_patents.method_name),\n"
            "  patent_type = CASE WHEN excluded.patent_type = '미분류'\n"
            "                     THEN COALESCE(pour_patents.patent_type, excluded.patent_type)\n"
            "                     ELSE excluded.patent_type END,\n"
            "  first_seen_at = COALESCE(NULLIF(excluded.first_seen_at, ''), pour_patents.first_seen_at),\n"
            "  last_seen_at = COALESCE(NULLIF(excluded.last_seen_at, ''), pour_patents.last_seen_at),\n"
            "  affiliation_type = CASE WHEN excluded.affiliation_type = '미분류'\n"
            "                          THEN COALESCE(pour_patents.affiliation_type, excluded.affiliation_type)\n"
            "                          ELSE excluded.affiliation_type END;"
        )
    (BASE / "nextjs" / "drizzle" / "seed-patent-master.sql").write_text(
        "\n".join(lines) + "\n", encoding="utf-8")

    # 원본 엑셀의 "구분" 열과 견주어 본다. 그 열은 현장 전체의 구분이므로 개별 특허의
    # 공법과 곧바로 같지는 않지만, 특허 한 건짜리 현장에서는 같아야 한다.
    # 어긋나는 것이 있으면 조용히 넘기지 않고 알린다.
    single = {}
    per_site = {}
    for row in sites:
        per_site.setdefault(row["client"], set()).add(row["number"])
    disagree = []
    for row in sites:
        if len(per_site[row["client"]]) != 1:
            continue
        raw = row["siteClassRaw"]
        if raw not in (TYPE_DO, TYPE_CNC, TYPE_POUR):
            continue
        got = master[row["number"]]["patentType"]
        if got != raw:
            disagree.append((row["number"], row["client"], raw, got))
        single[row["number"]] = raw

    by_type = {}
    for rec in records:
        by_type[rec["patentType"]] = by_type.get(rec["patentType"], 0) + 1
    site_names = {row["client"] for row in sites}
    multi = {}
    for row in sites:
        multi.setdefault(row["client"], set()).add(row["number"])
    multi_count = sum(1 for v in multi.values() if len(v) > 1)

    by_aff = {}
    for rec in records:
        by_aff[rec["affiliationType"]] = by_aff.get(rec["affiliationType"], 0) + 1
    own_total = sum(by_type.get(t, 0) for t in (TYPE_POUR, TYPE_DO, TYPE_CNC))

    print(f"특허 마스터        {len(records)}건 "
          f"(「특허(N)」 {from_master}건 + 현장에서 새로 본 {len(records) - from_master}건)")
    print("  공법별")
    for t in (TYPE_POUR, TYPE_DO, TYPE_CNC, TYPE_THIRD, TYPE_UNKNOWN):
        print(f"    {t:<6} {by_type.get(t, 0)}건")
    print("  소속별")
    for a in (AFF_OWN, AFF_THIRD, AFF_UNKNOWN):
        print(f"    {a:<6} {by_aff.get(a, 0)}건")
    print(f"  자사계열 전체 (POUR+DO+CNC)  {own_total}건")

    if filled:
        print(f"\n  특허 한 건짜리 현장의 '구분' 으로 채운 건 {len(filled)}개")
        for number, client, raw in filled:
            print(f"      {number} / {client} → {raw}")

    # 구분이 POUR 인데 업체가 DO·CNC 인 것은 옛 묶음이다. 이번에 나눈 것이 맞다.
    old_lumped = [d for d in disagree if d[2] == TYPE_POUR and d[3] in (TYPE_DO, TYPE_CNC)]
    other = [d for d in disagree if d not in old_lumped]
    if old_lumped:
        seen_nums = sorted({d[0] for d in old_lumped})
        print(f"\n  원본 '구분' 이 POUR 인데 업체가 DO·CNC 인 건 {len(old_lumped)}줄 "
              f"(특허번호 {len(seen_nums)}개) — 자사 계열을 POUR 로 묶어 세던 옛 방식입니다.")
        print("  이번 기준대로 DO·CNC 로 나눴습니다 (POUR 로 합치지 않았습니다).")
        for number, client, raw, got in old_lumped[:6]:
            print(f"      {number} / {client} — 원본:{raw} · 이번:{got}")
    if other:
        print(f"\n  ⚠ 그 밖에 어긋나는 건 {len(other)}개 (직접 확인이 필요합니다)")
        for number, client, raw, got in other[:10]:
            print(f"      {number} / {client} — 원본:{raw} · 판정:{got}")
    if not disagree:
        print("\n  원본 '구분' 열(특허 한 건짜리 현장)과 어긋나는 건 없음")
    print(f"현장별 특허 기록   {len(sites)}줄 · 현장 {len(site_names)}곳 · 특허 2개 이상 {multi_count}곳")
    print(f"\n  {out_json.relative_to(BASE.parent)}")
    print(f"  {(BASE / 'nextjs' / 'drizzle' / 'seed-patent-master.sql').relative_to(BASE.parent)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
